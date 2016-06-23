
import openpyxl
import xlrd
import csv

def csv_from_excel():
    """from so, check"""

    wb = xlrd.open_workbook('your_workbook.xls')
    sh = wb.sheet_by_name('Sheet1')
    your_csv_file = open('your_csv_file.csv', 'wb')
    wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

    for rownum in range(sh.nrows):
        wr.writerow(sh.row_values(rownum))

    your_csv_file.close()

def remove_double_classes():

    find_mergecells()

    unmerge_double_classes()


def clean_timetable():

    wb = openpyxl.load_workbook('B0.02 B0.03 B0.04 Timetable.xlsx')

    sheet_separator(wb)

    remove_double_classes()






# open the xlsx file to be operated on
wb = openpyxl.load_workbook('B0.02 B0.03 B0.04 Timetable.xlsx')


# make list of all rooms in the timetable sheet, excluding 'all'
# for future files, may be necessary to create a list of common exclusion sheets
# such as 'if i in excluded:'

rooms = []

for i in wb.get_sheet_names():
    if i == 'All':
        continue

    rooms.append(i)



# pull the first one to operate on

sheet = wb.get_sheet_by_name(rooms[0])



# print the value of a cell from the merged range list

sheet[sorted(sheet.merged_cell_ranges)[0][0:2]].value



# find first value in each cell merge range

mergecells = []

for i in range(len(sheet.merged_cell_ranges)):

    range_of_merge = sorted(sheet.merged_cell_ranges)[i]

    # only top left cell in merged cell contains a value, find this by splitting on ':' for 'A1:A2' = 'A1'
    top_left_cell_value = range_of_merge.split(":")[0]

    # several cells contain nothing but are merged due to format, disregard these.
    if sheet[top_left_cell_value].value == None:
        continue

    # create list of all the value-containing cells which begin merge cells
    mergecells.append(top_left_cell_value)



# only keep mergecells from the 'module' rows

# use row offset to remove the top two rows as these contain no classes
for row in sheet.iter_rows(row_offset=2):
    for cell in row:
        if cell.coordinate in mergecells:

            # assignment just to clean up function
            cell = cell.coordinate

            # create the cell range so 'A1' becomes 'A1:A2'
            cell_range = str(cell + ':' + cell[:1] ) + str( int(cell[1:]) + 1)

            sheet.unmerge_cells(cell_range)

            # copy value of first cell into the cell below it
            bottom_cell = cell_range.split(":")[1]
            sheet[bottom_cell].value = sheet[cell].value



wb.save('B002_finalmerge.xlsx')
